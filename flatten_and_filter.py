import pandas as pd
import re
from dateutil import parser as dateutil_parser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_field_simple(text, key):
    """Extract single-line field value."""
    if not text or pd.isna(text):
        return ""
    text = str(text)
    pattern = rf"(?:^|\n)\s*{re.escape(key)}:\s*(.+?)(?=\n|$)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return ""

def extract_multiline_field(text, key, stop_patterns):
    """Extract a field that may span multiple lines until a stop pattern."""
    if not text or pd.isna(text):
        return ""
    text = str(text)
    escaped_key = re.escape(key)
    stop = "|".join(re.escape(p) for p in stop_patterns)
    pattern = rf"{escaped_key}:\s*(.*?)(?=\n\s*(?:{stop}):|---|\Z)"
    match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
    if match:
        val = match.group(1).strip()
        val = re.sub(r'\s*\n\s*', ' ', val)
        return val.strip()
    return ""

def normalize_date(raw):
    """Parse any date string to YYYYMMDD; return raw string if unparseable."""
    if not raw:
        return ""
    # Remove ordinal suffixes (17th -> 17, 1st -> 1, etc.)
    cleaned = re.sub(r'(\d+)(st|nd|rd|th)\b', r'\1', raw, flags=re.IGNORECASE)
    # Strip time and timezone noise after | or bare time pattern
    cleaned = re.sub(r'\s*\|.*$', '', cleaned)
    cleaned = re.sub(r'\s+\d{1,2}:\d{2}(:\d{2})?\s*[A-Z]{2,3}$', '', cleaned)
    # Strip leading weekday abbreviations/names (e.g. "Fri. ", "Friday, ")
    cleaned = re.sub(r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-z.]*[,.]?\s*', '', cleaned, flags=re.IGNORECASE)
    try:
        dt = dateutil_parser.parse(cleaned, fuzzy=True, default=dateutil_parser.parse("2000-01-01"))
        return dt.strftime("%Y%m%d")
    except Exception:
        return raw

def parse_basic_info(text):
    raw_date = extract_field_simple(text, "Date")
    return {
        "Institution": extract_field_simple(text, "Institution"),
        "Title": extract_field_simple(text, "Title"),
        "Date": normalize_date(raw_date),
    }

def parse_view_section(text):
    if not text or pd.isna(text):
        return {k: "" for k in ["Classification", "Narrative", "Comparison Type", "EM Coverage",
                                  "Conviction Strength", "Conviction Direction", "Time Horizon",
                                  "Thesis Type", "Overall Explanation"]}
    text = str(text)
    single_line_keys = ["Classification", "Narrative", "Comparison Type", "EM Coverage",
                        "Conviction Strength", "Conviction Direction", "Time Horizon", "Thesis Type"]
    result = {}
    for k in single_line_keys:
        result[k] = extract_field_simple(text, k)
    result["Overall Explanation"] = extract_multiline_field(text, "Overall Explanation",
        ["Classification", "Narrative", "Comparison Type", "EM Coverage", "Conviction Strength",
         "Conviction Direction", "Time Horizon", "Thesis Type"])
    return result

def extract_priced_in(text):
    if not text or pd.isna(text):
        return ""
    text = str(text)
    match = re.search(r"Priced-In Expectations:\s*(.*?)$", text, re.DOTALL | re.IGNORECASE)
    if match:
        val = match.group(1).strip()
        val = re.sub(r'\s*\n\s*', ' ', val)
        return val.strip()
    return ""

def parse_drivers_section(text):
    if not text or pd.isna(text):
        return {}
    text = str(text)
    result = {}

    for field in ["Category", "Direction", "Explanation", "Evidence"]:
        result[f"Primary Driver {field}"] = extract_field_simple(text, f"Primary Driver {field}")

    for n in range(1, 4):
        for field in ["Category", "Direction", "Explanation", "Evidence"]:
            key = f"Supporting Driver {n} {field}"
            result[key] = extract_field_simple(text, key)

    for n in range(1, 4):
        result[f"Risk {n} Category"] = extract_field_simple(text, f"Risk {n} Category")
        result[f"Risk {n} Description"] = extract_multiline_field(
            text, f"Risk {n} Description",
            [f"Risk {n+1} Category", "Catalyst 1 Category", "Priced-In Expectations"])

    for n in range(1, 4):
        result[f"Catalyst {n} Category"] = extract_field_simple(text, f"Catalyst {n} Category")
        result[f"Catalyst {n} Description"] = extract_multiline_field(
            text, f"Catalyst {n} Description",
            [f"Catalyst {n+1} Category", "Priced-In Expectations"])

    result["Priced-In Expectations"] = extract_priced_in(text)
    return result


# Load raw data
df_raw = pd.read_excel(
    r"R:\Economics\Ziyan\PYTHON\EMvsUS Research\External Research Summary - EM vs US V2.xlsx",
    sheet_name=0, header=None
)

# Parse all rows (skip header row 0)
records = []
for idx in range(1, len(df_raw)):
    row = df_raw.iloc[idx]
    doc = str(row[0]) if pd.notna(row[0]) else ""
    record = {"Document": doc}
    record.update(parse_basic_info(row[1]))
    record.update(parse_view_section(row[2]))
    record.update(parse_drivers_section(row[3]))
    records.append(record)

df = pd.DataFrame(records)

# Filter: exclude "no explicit view" and "NA" classifications
mask = ~df["Classification"].str.strip().str.lower().isin(["no explicit view", "na", ""])
df_filtered = df[mask].reset_index(drop=True)

print(f"Total rows before filter: {len(df)}")
print(f"Total rows after filter:  {len(df_filtered)}")
print(f"\nClassification breakdown:\n{df_filtered['Classification'].value_counts()}")

# Write to Excel
wb = Workbook()
ws = wb.active
ws.title = "Filtered Data"

header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Arial", size=9)
data_align = Alignment(vertical="top", wrap_text=True)
fill_even = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
fill_odd = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
thin_border = Border(bottom=Side(style="thin", color="BDD7EE"))

cols = list(df_filtered.columns)
for col_idx, col_name in enumerate(cols, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for row_idx, (_, row_data) in enumerate(df_filtered.iterrows(), 2):
    fill = fill_even if row_idx % 2 == 0 else fill_odd
    for col_idx, col_name in enumerate(cols, 1):
        val = row_data[col_name]
        cell = ws.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else "")
        cell.font = data_font
        cell.alignment = data_align
        cell.fill = fill
        cell.border = thin_border

narrow_cols = {"Document": 20, "Institution": 20, "Title": 35, "Date": 18,
               "Classification": 22, "Narrative": 22, "Comparison Type": 18,
               "EM Coverage": 14, "Conviction Strength": 18, "Conviction Direction": 18,
               "Time Horizon": 14, "Thesis Type": 14}
wide_cols = {"Overall Explanation", "Primary Driver Explanation", "Primary Driver Evidence",
             "Risk 1 Description", "Risk 2 Description", "Risk 3 Description",
             "Catalyst 1 Description", "Catalyst 2 Description", "Catalyst 3 Description",
             "Priced-In Expectations"}
for n in range(1, 4):
    wide_cols.add(f"Supporting Driver {n} Explanation")
    wide_cols.add(f"Supporting Driver {n} Evidence")

for col_idx, col_name in enumerate(cols, 1):
    col_letter = get_column_letter(col_idx)
    if col_name in narrow_cols:
        ws.column_dimensions[col_letter].width = narrow_cols[col_name]
    elif col_name in wide_cols:
        ws.column_dimensions[col_letter].width = 55
    else:
        ws.column_dimensions[col_letter].width = 20

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

output_path = r"R:\Economics\Ziyan\PYTHON\EMvsUS Research\External Research Summary - EM vs US V2 (Filtered).xlsx"
wb.save(output_path)
print(f"\nSaved to: {output_path}")
