"""
Flatten and filter External Research - US_EAFE_EM.xlsx
Input : External Research - US_EAFE_EM.xlsx  (5 raw columns, multi-line cells)
Output: External Research - US_EAFE_EM (Filtered).xlsx  (one flat row per doc)
"""
import re
import pandas as pd
from dateutil import parser as dateutil_parser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE_DIR    = Path(__file__).parent
INPUT_FILE  = BASE_DIR / "External Research - US_EAFE_EM.xlsx"
OUTPUT_FILE = BASE_DIR / "External Research - US_EAFE_EM (Filtered).xlsx"

# ── Field extraction helpers ───────────────────────────────────────────────────

def field(text: str, key: str) -> str:
    """Extract a single-line value after 'key:'."""
    if not text or pd.isna(text):
        return ""
    m = re.search(
        rf"(?:^|\n)\s*{re.escape(key)}:\s*(.+?)(?=\n|$)",
        str(text), re.IGNORECASE
    )
    return m.group(1).strip() if m else ""


def multiline_field(text: str, key: str, stop_keys: list[str]) -> str:
    """Extract a value that may span multiple lines, stopping at the next key."""
    if not text or pd.isna(text):
        return ""
    stop = "|".join(re.escape(k) for k in stop_keys)
    m = re.search(
        rf"{re.escape(key)}:\s*(.*?)(?=\n\s*(?:{stop}):|---|\Z)",
        str(text), re.DOTALL | re.IGNORECASE
    )
    if not m:
        return ""
    return re.sub(r'\s*\n\s*', ' ', m.group(1)).strip()


def normalize_date(raw: str) -> str:
    if not raw:
        return ""
    cleaned = re.sub(r'(\d+)(st|nd|rd|th)\b', r'\1', raw, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s*\|.*$', '', cleaned)
    cleaned = re.sub(r'\s+\d{1,2}:\d{2}(:\d{2})?\s*[A-Z]{2,3}$', '', cleaned)
    cleaned = re.sub(r'^(Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-z.]*[,.]?\s*', '', cleaned, flags=re.IGNORECASE)
    try:
        dt = dateutil_parser.parse(cleaned, fuzzy=True, default=dateutil_parser.parse("2000-01-01"))
        return dt.strftime("%Y%m%d")
    except Exception:
        return raw


# ── Per-column parsers ─────────────────────────────────────────────────────────

def parse_short_summary(text: str) -> dict:
    return {
        "Doc Type":       field(text, "Document Type"),
        "Relevant Parties": field(text, "Relevant Parties"),
        "Relevant Dates": field(text, "Relevant Dates"),
    }


def parse_basic_info(text: str) -> dict:
    raw_date = field(text, "Date")
    return {
        "Institution": field(text, "Institution"),
        "Title":       field(text, "Title"),
        "Date":        normalize_date(raw_date),
    }


REGIONS = ["US", "EAFE", "EM"]
REGION_SINGLE = ["Classification", "Conviction Strength", "Time Horizon", "Thesis Type", "Narrative"]
REGION_MULTI  = ["Regional Summary"]

RELATIVE_KEYS = ["EM vs US", "EAFE vs US", "EM vs EAFE", "Relative Ranking",
                 "Dominant Global Theme", "Theme Explanation"]


def parse_positioning(text: str) -> dict:
    out: dict = {}
    if not text or pd.isna(text):
        for r in REGIONS:
            for k in REGION_SINGLE + REGION_MULTI:
                out[f"{r} {k}"] = ""
        for k in RELATIVE_KEYS:
            out[k] = ""
        return out

    for r in REGIONS:
        for k in REGION_SINGLE:
            out[f"{r} {k}"] = field(text, f"{r} {k}")
        # Regional Summary spans multiple lines
        stop = [f"{r2} Classification" for r2 in REGIONS if r2 != r] + \
               ["Relative Ranking", "EM vs US", "EAFE vs US", "EM vs EAFE",
                "Dominant Global Theme"]
        out[f"{r} Regional Summary"] = multiline_field(text, f"{r} Regional Summary", stop)

    for k in RELATIVE_KEYS:
        out[k] = field(text, k) if k != "Theme Explanation" else \
                 multiline_field(text, k, ["---"])

    return out


DRIVER_STOP = lambda n: [
    f"Driver {n+1} Name", "Risk 1 Primary Affected Region", "---"
]
RISK_STOP = lambda n: [
    f"Risk {n+1} Primary Affected Region", "Catalyst 1 Primary Affected Region", "---"
]
CAT_STOP = lambda n: [
    f"Catalyst {n+1} Primary Affected Region", "US Priced In", "EAFE Priced In",
    "EM Priced In", "---"
]


def parse_drivers(text: str) -> dict:
    out: dict = {}
    if not text or pd.isna(text):
        for n in range(1, 4):
            for f_ in ["Name", "Category", "Direction", "Explanation", "Evidence"]:
                out[f"Driver {n} {f_}"] = ""
            out[f"Risk {n} Primary Affected Region"] = ""
            out[f"Risk {n} Category"] = ""
            out[f"Risk {n} Description"] = ""
            out[f"Catalyst {n} Primary Affected Region"] = ""
            out[f"Catalyst {n} Category"] = ""
            out[f"Catalyst {n} Description"] = ""
        for r in REGIONS:
            out[f"{r} Priced In"] = ""
        return out

    for n in range(1, 4):
        out[f"Driver {n} Name"]      = field(text, f"Driver {n} Name")
        out[f"Driver {n} Category"]  = field(text, f"Driver {n} Category")
        out[f"Driver {n} Direction"] = field(text, f"Driver {n} Direction")
        out[f"Driver {n} Explanation"] = multiline_field(
            text, f"Driver {n} Explanation",
            [f"Driver {n} Evidence", f"Driver {n+1} Name",
             "Risk 1 Primary Affected Region", "---"])
        out[f"Driver {n} Evidence"] = multiline_field(
            text, f"Driver {n} Evidence",
            [f"Driver {n+1} Name", "Risk 1 Primary Affected Region", "---"])

    for n in range(1, 4):
        out[f"Risk {n} Primary Affected Region"] = field(text, f"Risk {n} Primary Affected Region")
        out[f"Risk {n} Category"]    = field(text, f"Risk {n} Category")
        out[f"Risk {n} Description"] = multiline_field(
            text, f"Risk {n} Description",
            [f"Risk {n+1} Primary Affected Region", "Catalyst 1 Primary Affected Region",
             "US Priced In", "---"])

    for n in range(1, 4):
        out[f"Catalyst {n} Primary Affected Region"] = field(text, f"Catalyst {n} Primary Affected Region")
        out[f"Catalyst {n} Category"]    = field(text, f"Catalyst {n} Category")
        out[f"Catalyst {n} Description"] = multiline_field(
            text, f"Catalyst {n} Description",
            [f"Catalyst {n+1} Primary Affected Region", "US Priced In",
             "EAFE Priced In", "EM Priced In", "---"])

    for r in REGIONS:
        out[f"{r} Priced In"] = multiline_field(
            text, f"{r} Priced In",
            [f"{r2} Priced In" for r2 in REGIONS if r2 != r] + ["---"])

    return out


# ── Main ───────────────────────────────────────────────────────────────────────

df_raw = pd.read_excel(INPUT_FILE, sheet_name=0, header=None)

records = []
for idx in range(1, len(df_raw)):
    row = df_raw.iloc[idx]
    doc = str(row[0]) if pd.notna(row[0]) else ""
    rec = {"Document": doc}
    rec.update(parse_short_summary(row[1]))
    rec.update(parse_basic_info(row[2]))
    rec.update(parse_positioning(row[3]))
    rec.update(parse_drivers(row[4]))
    records.append(rec)

df = pd.DataFrame(records)

# Filter: keep rows where at least two regions have a real classification
def has_classification(row):
    count = 0
    for r in REGIONS:
        val = str(row.get(f"{r} Classification", "")).strip().lower()
        if val and val not in ("not discussed", "na", "nan", ""):
            count += 1
    return count >= 2

mask = df.apply(has_classification, axis=1)
df_filtered = df[mask].reset_index(drop=True)

print(f"Total rows parsed:        {len(df)}")
print(f"Rows after filter:        {len(df_filtered)}")
for r in REGIONS:
    print(f"\n{r} Classification breakdown:")
    print(df_filtered[f"{r} Classification"].value_counts().to_string())

# ── Write styled Excel ─────────────────────────────────────────────────────────

wb = Workbook()
ws = wb.active
ws.title = "Filtered Data"

header_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font    = Font(name="Arial", size=9)
data_align   = Alignment(vertical="top", wrap_text=True)
fill_even    = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
fill_odd     = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
thin_border  = Border(bottom=Side(style="thin", color="BDD7EE"))

cols = list(df_filtered.columns)
for ci, name in enumerate(cols, 1):
    cell = ws.cell(row=1, column=ci, value=name)
    cell.font, cell.fill, cell.alignment = header_font, header_fill, header_align

for ri, (_, row_data) in enumerate(df_filtered.iterrows(), 2):
    fill = fill_even if ri % 2 == 0 else fill_odd
    for ci, name in enumerate(cols, 1):
        val = row_data[name]
        cell = ws.cell(row=ri, column=ci, value="" if pd.isna(val) else val)
        cell.font, cell.alignment, cell.fill, cell.border = \
            data_font, data_align, fill, thin_border

NARROW = {
    "Document": 22, "Doc Type": 18, "Relevant Parties": 20, "Relevant Dates": 16,
    "Institution": 20, "Title": 35, "Date": 12,
    "US Classification": 20, "EAFE Classification": 20, "EM Classification": 20,
    "US Conviction Strength": 18, "EAFE Conviction Strength": 18, "EM Conviction Strength": 18,
    "US Time Horizon": 14, "EAFE Time Horizon": 14, "EM Time Horizon": 14,
    "US Thesis Type": 14, "EAFE Thesis Type": 14, "EM Thesis Type": 14,
    "US Narrative": 18, "EAFE Narrative": 18, "EM Narrative": 18,
    "EM vs US": 16, "EAFE vs US": 16, "EM vs EAFE": 16, "Relative Ranking": 16,
    "Dominant Global Theme": 20,
}
WIDE = {
    "US Regional Summary", "EAFE Regional Summary", "EM Regional Summary",
    "Theme Explanation",
    "Driver 1 Explanation", "Driver 1 Evidence",
    "Driver 2 Explanation", "Driver 2 Evidence",
    "Driver 3 Explanation", "Driver 3 Evidence",
    "Risk 1 Description", "Risk 2 Description", "Risk 3 Description",
    "Catalyst 1 Description", "Catalyst 2 Description", "Catalyst 3 Description",
    "US Priced In", "EAFE Priced In", "EM Priced In",
}

for ci, name in enumerate(cols, 1):
    letter = get_column_letter(ci)
    if name in NARROW:
        ws.column_dimensions[letter].width = NARROW[name]
    elif name in WIDE:
        ws.column_dimensions[letter].width = 55
    else:
        ws.column_dimensions[letter].width = 20

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_FILE)
print(f"\nSaved → {OUTPUT_FILE}")
